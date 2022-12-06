import sys
import os
import csv
from openpyxl import load_workbook


class HD1CodePlugSystem:

    def __init__(self,
                 name, system_sheet_name, tx, rx, template, systype, radio_id
                ):
        self._name = name
        self._system_sheet_name = system_sheet_name
        self._tx = tx
        self._rx = rx
        self._template = template
        self._type = systype
        self._radio_id = radio_id

        self._talkgroups = {}
        self._channels = {}

    def add_talkgroup(self, talkgroup):
        self._talkgroups[talkgroup._talkgroup] = talkgroup

    def add_channel(self, channel):
        self._channels[channel._channel_alias] = channel

    def talkgroups(self):
        return self._talkgroups

    def channels(self):
        return self._channels


class HD1CodePlugTemplate:
    
    def __init__(self,
                 name, data
                ):
        self._name = name
        self._data = data


class HD1CodePlugChannel:

    def __init__(self, 
                 system
                ):
        self._system = system
        self._template = None
        self._number = 0

    def _replace(self, working, name, value):
        index = 0
        for cell in working:
            if cell == name:
                working[index] = value
            index = index + 1

    def create_template_fields(self):
        return self._template._data.copy()

    def populate_fields(self, fields):
        if self._number != 0:
            self._replace(fields, "$NUMBER", self._number)
        self._replace(fields, "$ALIAS", self._channel_alias)
        self._replace(fields, "$RADIOID", self._system._radio_id)


class HD1CodePlugTalkGroupChannel(HD1CodePlugChannel):
    
    def __init__(self,
                 system,
                 talkgroup, 
                 slot, 
                 name, 
                 channel_alias
                ):
        HD1CodePlugChannel.__init__(self, system)

        self._talkgroup = talkgroup
        self._slot = slot
        self._name = name
        self._channel_alias = channel_alias

    def populate_fields(self, fields):
        super().populate_fields(fields)

        self._replace(fields, "$SLOT", "Slot{0}".format(self._slot))
        self._replace(fields, "$TX", self._system._tx)
        self._replace(fields, "$RX", self._system._rx)
        self._replace(fields, "$CONTACT", "Priority Contacts: TG {0}".format(self._talkgroup))


class HD1CodePlugFrequencyChannel(HD1CodePlugChannel):

    def __init__(self,
                 system,
                 channel_alias,
                 rx,
                 tx,
                 radioid):
        HD1CodePlugChannel.__init__(self, system)

        self._channel_alias = channel_alias 
        self._rx = rx
        self._tx = tx
        self._radio_id = radioid

    def populate_fields(self, fields):
        super().populate_fields(fields)

        self._replace(fields, "$TX", self._tx)
        self._replace(fields, "$RX", self._rx)


class HD1CodePlugVFOChannel(HD1CodePlugFrequencyChannel):

    def __init__(self,
                 system,
                 name,
                 channel_alias,
                 rx,
                 tx,
                 radioid):
        HD1CodePlugFrequencyChannel.__init__(self, system, channel_alias, rx, tx, radioid)

        self._name = name

    def populate_fields(self, working):
        super().populate_fields(working)

        self._replace(working, "$NAME", self._name)


class HD1CodePlugAnalogRepaterChannel(HD1CodePlugFrequencyChannel):
    
    def __init__ (self, system, key, tx, rx, ctcss, radio_id):
        HD1CodePlugFrequencyChannel.__init__(self, system, key, rx, tx, radio_id)
        self._ctcss = ctcss

    def populate_fields(self, working):
        super().populate_fields(working)

        self._replace(working, "$CTCSS", self._ctcss)

class HD1CodePlugDigitalRepeaterChannel(HD1CodePlugFrequencyChannel):
    
    def __init__ (self, system, key, tx, rx, colour, radio_id):
        HD1CodePlugFrequencyChannel.__init__(self, system, key, rx, tx, radio_id)
        self._colour = colour

    def populate_fields(self, working):
        super().populate_fields(working)

        self._replace(working, "$COLOUR", self._colour)


class HD1CodePlugPriorityContact:

    def __init__(self,
                 number,
                 call_type,
                 contact_alias,
                 call_id
                ):

        self._number = number
        self._call_type = call_type
        self._contact_alias = contact_alias
        self._call_id = call_id

    def __str__(self):
        return "{0}, {1}, {2}, {3}".format(self._number, self._call_type, self._contact_alias, self._call_id)

    def populate_fields(self):
        return [self._number, self._call_type, self._call_type, "", "", "", self._call_id]


BASE_INFO_SHEET = "HD1 Base Info"

PRIORITY_CONTACTS_SHEET = "HD1 Priority Contacts"
PRIORITY_CONTACTS_HEADER = "No.,Call Type,Contacts Alias,City,Province,Country,Call ID"

CHANNEL_INFORMATION_SHEET = "HD1 Channel Information"
CHANNEL_INFORMATION_HEADER = "No.,Channel Type,Channel Alias,Rx Frequency,Tx Frequency,Tx Power,TOT,VOX,VOX Level,Scan Add/Step,Channel Work Alone,Default to Talkaround,Band Width,Dec QT/DQT,Enc QT/DQT,Tx Authority,Relay,Work Mode,Slot,ID Setting,Color Code,Encryption,Encryption Type,Encryption Key,Promiscuous,Tx Authority,Kill Code,WakeUp Code,Contacts,Rx Group Lists,Group Lists 1,Group Lists 2,Group Lists 3,Group Lists 4,Group Lists 5,Group Lists 6,Group Lists 7,Group Lists 8,Group Lists 9,Group Lists 10,Group Lists 11,Group Lists 12,Group Lists 13,Group Lists 14,Group Lists 15,Group Lists 16,Group Lists 17,Group Lists 18,Group Lists 19,Group Lists 20,Group Lists 21,Group Lists 22,Group Lists 23,Group Lists 24,Group Lists 25,Group Lists 26,Group Lists 27,Group Lists 28,Group Lists 29,Group Lists 30,Group Lists 31,Group Lists 32,Group Lists 33,GPS,Send GPS Info,Receive GPS Info,GPS Timing Report,GPS Timing Report TX Contacts"

ADDRESS_BOOK_CONTACTS_SHEET = "HD1 Address Book Contacts"

VFO_CHANNEL_INFO_SHEET = "VFO Channel Info"

ZONE_INFOMATION_SHEET = "HD1 Zone Information"
ZONE_INFOMATION_HEADER = "System, Start, End"


class HD1CodePlugSpreadsheet:

    def __init__(self, 
                 spreadsheet_filename,
                 config_sheet = BASE_INFO_SHEET):
        
        print("Loading workbook: ", spreadsheet_filename)

        self._spreadsheet_filename = spreadsheet_filename
        self._config_sheet = config_sheet

        self._workbook = load_workbook(self._spreadsheet_filename, data_only=True)

        self._radio_id = "RADIOID"

        self._systems = {}
        self._templates = {}

        self._priority_contacts = []
        self._channels = []

        self._config = self.load_config_sheet()

        self._vfo = HD1CodePlugSystem('VFO', VFO_CHANNEL_INFO_SHEET, None, None, None, 'Channels', self._radio_id)

    def load_config_sheet(self):

        if self._config_sheet not in self._workbook:
            raise Exception("Config sheet '{0}' missing or mispelled".format(self._config_sheet))

        print("Loading configuration from: ", self._config_sheet)

        self._load_base_info()

    def _load_base_info(self):
        base_info = self._workbook[self._config_sheet]
        self._load_radio_id(base_info)
        self._load_system_config(base_info)
        self._load_templates(base_info)

    def _find_start_row(self, sheet, text):
        line = 1
        for row in sheet:
            if row[0].value == text:
                return line
            line = line + 1

        return -1

    def _load_radio_id(self, base_info):
        line_count = self._find_start_row(base_info, 'Radio ID')
        if line_count == -1:
            raise Exception ("No Radio ID defined in config") 
        line_count = line_count + 1  

        self._radio_id = name = base_info["A{0}".format(line_count)].value

    def check_frequency(frequency):
        if 136.00 <= frequency <= 174.00:
            return True
        if 400.00 <= frequency <= 480.00:
            return True
        return False

    def _load_system_config(self, base_info):

        line_count = self._find_start_row(base_info, 'System')
        if line_count == -1:
            raise Exception ("No Systems defined in config") 
        line_count = line_count + 1  

        process = True
        while process is True:
            name = base_info["A{0}".format(line_count)].value
            if name is not None and name != '':
                include = base_info["B{0}".format(line_count)].value
                if include == 'Y':
                    system_sheet_name = base_info["C{0}".format(line_count)].value
                    tx = base_info["D{0}".format(line_count)].value
                    rx = base_info["E{0}".format(line_count)].value
                    template = base_info["F{0}".format(line_count)].value
                    systype = base_info["G{0}".format(line_count)].value

                    if tx is not None and HD1CodePlugSpreadsheet.check_frequency(tx) is False:
                        print("TX Frequency out of bounds for {0}".format(system_sheet_name))
                    elif rx is not None and HD1CodePlugSpreadsheet.check_frequency(rx) is False:
                        print("RX Frequency out of bounds for {0}".format(system_sheet_name))
                    else:
                        self._systems[name] = HD1CodePlugSystem(name, system_sheet_name, tx, rx, template, systype, self._radio_id)

            else:
                process = False  
            line_count = line_count + 1     

    def _load_templates(self, base_info):

        line_count = self._find_start_row(base_info, 'Templates')
        if line_count == -1:
            raise Exception ("No Templates defined in config") 
        line_count = line_count + 1  

        process = True
        while process is True:
            name = base_info["A{0}".format(line_count)].value
            if name is not None and name != '':
                end_column = base_info["B{0}".format(line_count)].value

                start_cell = "C{0}".format(line_count)
                end_cell = "{0}{1}".format(end_column, line_count)

                items = base_info["{0}".format(start_cell):"{0}".format(end_cell)]
                data = [cell.value for cell in items[0]]

                self._templates[name] = HD1CodePlugTemplate(name, data)

            else:
                process = False  

            line_count = line_count + 1     

    def load_systems(self):

        for system in self._systems.values():

            if system._system_sheet_name in self._workbook.sheetnames:
                print("Loading system for: {0} = {1}".format(system._system_sheet_name, system._type))

                worksheet = self._workbook[system._system_sheet_name]
                first = True
                for row in worksheet:
                    if first is False:
                        
                        if system._type == "Talkgroups":
                            talkgroup = row[0].value
                            slot = row[1].value
                            name = row[2].value
                            channel_alias = row[3].value

                            system.add_talkgroup(HD1CodePlugTalkGroupChannel(system, talkgroup, slot, name, channel_alias))

                        elif system._type == "Channels":
                            channel_alias = row[0].value
                            tx = row[1].value
                            rx = row[2].value

                            if tx is not None and HD1CodePlugSpreadsheet.check_frequency(tx) is False:
                                print("TX Frequency out of bounds for {0}".format(system._system_sheet_name, channel_alias))
                            elif rx is not None and HD1CodePlugSpreadsheet.check_frequency(rx) is False:
                                print("RX Frequency out of bounds for {0}".format(system._system_sheet_name, channel_alias))
                            else:
                                system.add_channel(HD1CodePlugFrequencyChannel(system, channel_alias, tx, rx, system._radio_id))

                        elif system._type == "ARepeaters":
                            key = row[0].value
                            tx = row[2].value
                            rx = row[3].value
                            ctcss = row[9].value

                            if tx is not None and HD1CodePlugSpreadsheet.check_frequency(tx) is False:
                                print("TX Frequency out of bounds for {0}".format(system._system_sheet_name, key))
                            elif rx is not None and HD1CodePlugSpreadsheet.check_frequency(rx) is False:
                                print("RX Frequency out of bounds for {0}".format(system._system_sheet_name, key))
                            else:
                                system.add_channel(HD1CodePlugAnalogRepaterChannel(system, key, tx, rx, ctcss, system._radio_id))

                        elif system._type == "DRepeaters":
                            key = row[0].value
                            tx = row[2].value
                            rx = row[3].value
                            colour = row[10].value

                            if tx is not None and HD1CodePlugSpreadsheet.check_frequency(tx) is False:
                                print("TX Frequency out of bounds for {0}".format(system._system_sheet_name, key))
                            elif rx is not None and HD1CodePlugSpreadsheet.check_frequency(rx) is False:
                                print("RX Frequency out of bounds for {0}".format(system._system_sheet_name, key))
                            else:
                                system.add_channel(HD1CodePlugDigitalRepeaterChannel(system, key, tx, rx, colour, system._radio_id))

                        else:
                            print("Unknown system type: ", system._type)

                    else:
                        first = False
            else:
                print("Missing talkgroup worksheet '{0}'".format(system._system_sheet_name))

    def load_vfo_channels(self):
        print ("Loading VFO Channels")

        if VFO_CHANNEL_INFO_SHEET in self._workbook.sheetnames:
            vfo_ws = self._workbook[VFO_CHANNEL_INFO_SHEET]
            first = True
            for row in vfo_ws:
                if first is False:
                    self._vfo._channels[row[0].value] = HD1CodePlugVFOChannel(self._vfo, row[0].value, row[1].value, row[2].value, row[3].value, self._radio_id)
                else:
                    first = False
        else:
            print("Missing cfo worksheet '{0}'".format(VFO_CHANNEL_INFO_SHEET))

    def create_priority_contacts(self):

        print("Generating Priority Contacts")

        talkgroup_ids = []

        zones = {}

        count = 1
        start = end = 0
        for system in self._systems.values():
            print("\tProcessing system for:", system._system_sheet_name)
            start = count
            for tg in system.talkgroups():
                system._talkgroups[tg]._number = count
                talkgroup_ids.append(tg)
                count = count + 1

            for ch in system.channels():
                system._channels[ch]._number = count
                count = count + 1

            end = count - 1
            zones[system._name] = [start, end]

        talkgroup_ids.sort()

        deduped_ids = list(dict.fromkeys(talkgroup_ids))

        count = 1
        for id in deduped_ids:
            pc = HD1CodePlugPriorityContact(count, "Group Call", "TG {0}".format(id), id)
            self._priority_contacts.append(pc)
            count = count + 1

        self._write_priority_contacts_to_worksheet()

        self._write_zones_to_worksheet(zones)

    def _create_new_worksheet(self, worksheet_name, idx):

        if worksheet_name in self._workbook.sheetnames:
            idx = self._workbook.sheetnames.index(worksheet_name)
            self._workbook.remove(self._workbook[worksheet_name])

        ws =  self._workbook.create_sheet(worksheet_name, idx)

        return ws

    def _write_priority_contacts_to_worksheet(self):

        priority_contacts_ws = self._create_new_worksheet(PRIORITY_CONTACTS_SHEET, 1)

        fields = PRIORITY_CONTACTS_HEADER.split(",")
        column = 1
        for field in fields:
            priority_contacts_ws.cell(1, column).value = field
            column = column + 1

        row = 2        
        for contact in self._priority_contacts:
            fields = contact.populate_fields()

            column = 1
            for field in fields:
                priority_contacts_ws.cell(row, column).value = field
                column = column + 1

            row = row + 1

    def _write_zones_to_worksheet(self, zones):

        zone_info_ws = self._create_new_worksheet(ZONE_INFOMATION_SHEET, 3)

        fields = ZONE_INFOMATION_HEADER.split(",")
        column = 1
        for field in fields:
            zone_info_ws.cell(1, column).value = field
            column = column + 1

        row = 2        
        for zone_name in zones.keys():

            zone = zones[zone_name]

            zone_info_ws.cell(row, 1).value = zone_name
            zone_info_ws.cell(row, 2).value = zone[0]
            zone_info_ws.cell(row, 3).value = zone[1]

            row = row + 1

    def create_channel_information(self):

        print("Generating Channel Informatio")

        self._create_vfo_channel_info()

        self._create_talkgroup_channel_info()

    def _create_vfo_channel_info(self):

        template = self._templates['VFO']

        for vfo in self._vfo._channels.values():
            vfo._template = template
            self._channels.append(vfo)

    def _create_talkgroup_channel_info(self):

        for system in self._systems.values():
            print("\tProcessing system for:", system._system_sheet_name)
   
            print("\t\tLoading template: ", system._template)
            template = self._templates[system._template]

            for tg in system._talkgroups.values():
                tg._template = template
                self._channels.append(tg)

            for ch in system._channels.values():
                ch._template = template
                self._channels.append(ch)

        self._write_channel_info_to_worksheet()

    def _write_channel_info_to_worksheet(self):

        channel_info_ws =  self._create_new_worksheet(CHANNEL_INFORMATION_SHEET, 2)

        self._write_channel_info_header_to_worksheet(channel_info_ws)

        self._write_channels_info_to_worksheet(channel_info_ws)

    def _write_channel_info_header_to_worksheet(self, channel_info_ws):

        fields = CHANNEL_INFORMATION_HEADER.split(",")
        column = 1
        for field in fields:
            channel_info_ws.cell(1, column).value = field
            column = column + 1

    def _write_channels_info_to_worksheet(self, channel_info_ws):
        row = 2

        for channel in self._channels:

            fields = channel.create_template_fields()
            channel.populate_fields(fields)

            column = 1
            for field in fields:
                channel_info_ws.cell(row, column).value = field
                column = column + 1

            row = row + 1

    def create_xlsx(self, output_filename = None):
        
        codeplug.load_systems()
        codeplug.load_vfo_channels()

        codeplug.create_priority_contacts()
        codeplug.create_channel_information()

        if output_filename is None:
            codeplug.save(self._spreadsheet_filename)
        else:
            codeplug.save(output_filename)

    def save(self, filename):
        print("Saving workbook '{0}".format(filename))

        self._workbook.save(filename)

    def create_csvs(self):

        self._export_sheet_csv(PRIORITY_CONTACTS_SHEET)
        self._export_sheet_csv(CHANNEL_INFORMATION_SHEET)
        self._export_sheet_csv(ADDRESS_BOOK_CONTACTS_SHEET)

    def _export_sheet_csv(self, sheet_name):

        if sheet_name in self._workbook.sheetnames:

            print("Writing {0} to csv".format(sheet_name))

            ws = self._workbook[sheet_name]
            if not ws:
                print("Worksheet '{0}' does not existing, skipping csv".format(sheet_name))

            filepath = sheet_name+".csv"

            if os.path.exists(filepath):
                os.remove(filepath)

            with open(filepath, 'w', newline="") as f:
                c = csv.writer(f)
                for r in ws.rows:
                    c.writerow([cell.value for cell in r])


if __name__ == '__main__':

    codeplug = HD1CodePlugSpreadsheet(sys.argv[1])
    
    if sys.argv[2] == 'xlsx':
        output_filename = None
        if len(sys.argv) == 4:
            output_filename = sys.argv[3]
        codeplug.create_xlsx(output_filename)

    elif sys.argv[2] == 'csv':
        codeplug.create_csvs()

    else:
        print("Unknown command line option xlsx or csv only")